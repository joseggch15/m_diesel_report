# -*- coding: utf-8 -*-
"""
Modelo de datos y logica de generacion del
"Monthly Diesel Tank reconciliation report".

Reporta UN producto (Diesel) y un tanque consolidado (LFO Main + Virtual),
mas tres service trucks (TFL0846, 0847, 0848).

Hermano de diesel_report.report_model con dos diferencias clave:
  - La ventana es UN MES calendario (primer dia -> ultimo dia).
  - Se agrega: weekly breakdown del mes, monthly trend (Delivery % + Recon %),
    historical Transaction Volume by Type y Tabla 1b con ~30 filas diarias.

No contiene interfaz grafica (ver app_qt.py).
"""
from __future__ import annotations

import calendar
import datetime
import os
import shutil
import tempfile
import zipfile

import openpyxl
from docxtpl import DocxTemplate, InlineImage
from docxtpl import RichText as _DocxRichText
from docx.shared import Mm

import charts


# --------------------------------------------------------------------------
# RichText con orden de propiedades valido segun el esquema OOXML
# --------------------------------------------------------------------------
#
# docxtpl 0.20.x emite las propiedades de run (<w:rPr>) en un orden que no
# respeta la secuencia obligatoria del esquema CT_RPr de OOXML: por ejemplo
# pone <w:b/> antes de <w:rFonts/>. El esquema exige que <w:rFonts/> vaya
# ANTES de <w:b/>, <w:color/>, <w:sz/>, etc. Word valida ese orden y, cuando
# no se cumple, muestra el dialogo "Word encontro contenido no legible... el
# desea recuperar el contenido?" al abrir el .docx.
#
# Esta subclase reescribe `add` para generar las propiedades en el orden que
# manda el esquema, eliminando ese mensaje. El resto del comportamiento es
# identico al de docxtpl.

try:
    from html import escape as _xml_escape
except ImportError:  # pragma: no cover
    from cgi import escape as _xml_escape


class RichText(_DocxRichText):
    """RichText de docxtpl con el orden de <w:rPr> corregido."""

    def add(
        self,
        text,
        style=None,
        color=None,
        highlight=None,
        size=None,
        subscript=None,
        superscript=None,
        bold=False,
        italic=False,
        underline=False,
        strike=False,
        font=None,
        url_id=None,
        rtl=False,
        lang=None,
    ):
        # Si se agrega otro RichText, concatenar su XML tal cual.
        if isinstance(text, _DocxRichText):
            self.xml += text.xml
            return

        if not isinstance(text, (str, bytes)):
            text = str(text)
        if not isinstance(text, str):
            text = text.decode("utf-8", errors="ignore")
        text = _xml_escape(text)

        # Las propiedades se emiten en el orden de la secuencia CT_RPr:
        # rStyle -> rFonts -> b/bCs -> i/iCs -> strike -> color -> sz/szCs
        # -> highlight(shd) -> u -> vertAlign -> rtl -> lang
        prop = ""
        if style:
            prop += '<w:rStyle w:val="%s"/>' % style
        if font:
            regional_font = ""
            if ":" in font:
                region, font = font.split(":", 1)
                regional_font = ' w:{region}="{font}"'.format(
                    font=font, region=region)
            prop += ('<w:rFonts w:ascii="{font}" w:hAnsi="{font}" '
                     'w:cs="{font}"{regional_font}/>').format(
                font=font, regional_font=regional_font)
        if bold:
            prop += "<w:b/>"
            if rtl:
                prop += "<w:bCs/>"
        if italic:
            prop += "<w:i/>"
            if rtl:
                prop += "<w:iCs/>"
        if strike:
            prop += "<w:strike/>"
        if color:
            if color[0] == "#":
                color = color[1:]
            prop += '<w:color w:val="%s"/>' % color
        if size:
            prop += '<w:sz w:val="%s"/>' % size
            prop += '<w:szCs w:val="%s"/>' % size
        if highlight:
            if highlight[0] == "#":
                highlight = highlight[1:]
            prop += '<w:shd w:fill="%s"/>' % highlight
        if underline:
            if underline not in [
                "single", "double", "thick", "dotted", "dash",
                "dotDash", "dotDotDash", "wave",
            ]:
                underline = "single"
            prop += '<w:u w:val="%s"/>' % underline
        if subscript:
            prop += '<w:vertAlign w:val="subscript"/>'
        if superscript:
            prop += '<w:vertAlign w:val="superscript"/>'
        if rtl:
            prop += '<w:rtl w:val="true"/>'
        if lang:
            prop += '<w:lang w:val="%s"/>' % lang

        xml = "<w:r>"
        if prop:
            xml += "<w:rPr>%s</w:rPr>" % prop
        xml += '<w:t xml:space="preserve">%s</w:t></w:r>' % text
        if url_id:
            xml = ('<w:hyperlink r:id="%s" w:tgtFrame="_blank">%s</w:hyperlink>'
                   % (url_id, xml))
        self.xml += xml

# --------------------------------------------------------------------------
# Configuracion fija
# --------------------------------------------------------------------------

PRODUCT_KEY = "Diesel"
SITE_KEY = "LFO"
SERVICE_TRUCKS = ["TFL0846", "TFL0847", "TFL0848"]

# Figuras del reporte
FIGURE_KEYS = [
    "fig1",        # Daily Inlet Deliveries vs Delivery Tickets (bars)
    "fig2",        # Fuel Outflow Distribution (donut)
    "fig3",        # Daily Diesel Tank Log
    "fig4",        # Reconciliation trend (monthly Recon %)
    "fig5",        # Delivery vs Reconciliation (Delivery % + Recon %)
    "fig6",        # Monthly Breakdown of Fuel Consumption (weekly bars)
    "fig7",        # Historical Transaction Volume by Type
    "fig_tasks",   # Imagen de la tabla de tareas (manual)
]

FIGURE_LABELS = {
    "fig1": "Figura 1 - Detailed Inlet Deliveries vs Delivery Tickets",
    "fig2": "Figura 2 - Fuel Outflow Distribution (donut)",
    "fig3": "Figura 3 - Daily Diesel Tank Log",
    "fig4": "Figura 4 - Reconciliation trend (mensual)",
    "fig5": "Figura 5 - Delivery vs Reconciliation",
    "fig6": "Figura 6 - Monthly Breakdown of Fuel Consumption",
    "fig7": "Figura 7 - Historical Transaction Volume by Type",
    "fig_tasks": "Tabla de tareas (imagen)",
}

# Anchos sugeridos (mm).
_FIG_WIDTHS_MM = {
    "fig1": 160,
    "fig2": 100,
    "fig3": 160,
    "fig4": 150,
    "fig5": 150,
    "fig6": 150,
    "fig7": 160,
    "fig_tasks": 160,
}


def figure_width(fig_key: str) -> Mm:
    return Mm(_FIG_WIDTHS_MM.get(fig_key, 150))


# Consideraciones del reporte mensual (4 bullets, distintos al weekly).
DEFAULT_CONSIDERATIONS = [
    "All fuel coming into the Storage tanks from Sol through the LFO Lane 1 "
    "inlet meters is categorized as Deliveries.",
    "All fuel dispensed from outlet meters or dispensers into equipment is "
    "categorized as Transaction.",
    "Opening stock corresponds to the sum of the Main Tank and Virtual Tank "
    "volumes at the date and time marking the start of the reconciliation "
    "period.",
    "Closing stock corresponds to the sum of the Main Tank and Virtual Tank "
    "volumes at the date and time marking the end of the reconciliation "
    "period.",
]

REPORT_FONT = "Arial"

# Umbrales editables para la narrativa de deliveries.
DEFAULT_BELOW_THRESHOLD_PCT = -0.50   # "below the -0,50% threshold"


# --------------------------------------------------------------------------
# Utilidades de numeros y formato
# --------------------------------------------------------------------------

def _num(value, default=0.0) -> float:
    if value is None or value == "":
        return float(default)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return float(default)


def fmt_int(value) -> str:
    n = round(_num(value))
    return "{:,}".format(int(n)).replace(",", ".")


def fmt_pct(value, decimals=2) -> str:
    text = ("{:." + str(decimals) + "f}").format(_num(value))
    return text.replace(".", ",") + "%"


_MONTHS_EN = ["January", "February", "March", "April", "May", "June", "July",
              "August", "September", "October", "November", "December"]
_MONTHS_EN_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug",
                    "Sep", "Oct", "Nov", "Dec"]


def ordinal(n: int) -> str:
    n = int(n)
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return "%d%s" % (n, suffix)


def format_month_label(d) -> str:
    """'April 2026'."""
    return "%s %d" % (_MONTHS_EN[d.month - 1], d.year)


def format_period_full(start, end) -> str:
    """'1st April to 30th 2026' (replica del DOCX manual)."""
    return "%s %s to %s %d" % (
        ordinal(start.day), _MONTHS_EN[start.month - 1],
        ordinal(end.day), end.year)


def format_cover_date(d) -> str:
    return "%s %s, %d" % (_MONTHS_EN[d.month - 1], ordinal(d.day), d.year)


def first_last_of_month(d) -> tuple:
    if isinstance(d, datetime.datetime):
        d = d.date()
    first = d.replace(day=1)
    last_day = calendar.monthrange(first.year, first.month)[1]
    last = first.replace(day=last_day)
    return first, last


# --------------------------------------------------------------------------
# Estructura de datos por defecto
# --------------------------------------------------------------------------

def default_data() -> dict:
    """Valores de ejemplo del reporte de Abril 2026."""
    return {
        "period_full": "1st April to 30th 2026",
        "month_label": "April 2026",
        "cover_date": "April 30th, 2026",
        # Tabla 1 (consolidada)
        "consolidated": {
            "site": SITE_KEY,
            "opening": 879250.31,
            "deliveries": 4263438.10,
            "transactions": 4214579,
            "closing": 920906.67,
        },
        # Tabla 1b: una fila por dia del mes
        # date como 'd-mmm' (string), tickets y inlet en L
        "deliveries_daily": [
            # {"date": "1-apr", "tickets": 154000, "inlet": 153327}
        ],
        # Tabla 2: Service Trucks (3)
        "service_trucks": [
            {"site": "TFL0846", "deliveries": 789604, "transactions": 814735},
            {"site": "TFL0847", "deliveries": 816310, "transactions": 830035},
            {"site": "TFL0848", "deliveries": 8039,   "transactions": 14471},
        ],
        # Dispensing (Figura 2)
        "dispensing": {
            "to_equipment": 2600625,
            "other": 0,
            "transfers": 1613953,
        },
        # Stats del Tank Log (parrafo bajo Figura 3)
        "tank_log_stats": {
            "min_value": 700000,
            "min_date": "9th April",
            "average": 900000,
        },
        # Monthly Transactions overview (texto bajo Figura 5)
        "monthly_overview": {
            "transactions_avg": 899784,
            "to_equipment_avg": 556250,
            "other_avg": 0,
            "transfers_avg": 343535,
        },
        # Tabla 3: Weekly breakdown del mes (Week | Date | Transactions | ToEq | Transfers)
        "weekly_breakdown": [
            # {"week": 1, "date": "06/04/26", "transactions": 1005485, "to_equipment": 643584, "transfers": 361901}
        ],
        # Datos para Fig 4 + 5 (puntos de Monthly Variance)
        "monthly_trend": [
            # (date, delivery_pct, recon_pct)
        ],
        # Historico mensual de outflows para Fig 7 (PVT OUTFLOW)
        "historical_outflows": [
            # (label, transactions, to_equipment, transfers, other)
        ],
        # Conteos para la narrativa de deliveries
        "deliveries_summary": {
            "confirmed_count": 0,
            "below_threshold_count": 0,
            "below_threshold_pct": DEFAULT_BELOW_THRESHOLD_PCT,
            "band_minus1_to_0_count": 0,
            "avg_variance_pct": 0.0,
            "min_variance_pct": 0.0,
            "max_variance_pct": 0.0,
            "avg_daily_volume": 0.0,
        },
        # Frase de variance editable (vacia = autogenerada)
        "recon_sentence": "",
        # Considerations (bullets)
        "considerations": list(DEFAULT_CONSIDERATIONS),
        # Narrativa de deliveries (vacia = autogenerada)
        "delivery_narrative": "",
        # Tasks (bullets bajo "Pending items" / "continuous improvement")
        "tasks_notes": [
            "Working on contractors' data to upload to AdaptIQ.",
            "Process to tag rental LVs.",
            "Process to install tags on Site Services Fleet",
            "Tag all contractor's fleet.",
        ],
    }


# --------------------------------------------------------------------------
# Lectura / escritura del Excel de trabajo (editable por el usuario)
# --------------------------------------------------------------------------

SHEET_META = "Meta"
SHEET_CONSOLIDATED = "Consolidated"
SHEET_TRUCKS = "ServiceTrucks"
SHEET_DELIVERIES = "DeliveriesDaily"
SHEET_DISPENSING = "Dispensing"
SHEET_WEEKLY = "WeeklyBreakdown"
SHEET_TREND = "MonthlyTrend"
SHEET_HIST = "HistoricalOutflows"
SHEET_TEXTS = "Texts"
SHEET_TASKS = "Tasks"


def save_excel(data: dict, path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_META
    ws.append(["field", "value"])
    ws.append(["period_full", data["period_full"]])
    ws.append(["month_label", data.get("month_label", "")])
    ws.append(["cover_date", data.get("cover_date", "")])
    ws.append(["tank_log_min_value", data["tank_log_stats"]["min_value"]])
    ws.append(["tank_log_min_date", data["tank_log_stats"]["min_date"]])
    ws.append(["tank_log_average", data["tank_log_stats"]["average"]])
    for k in ("transactions_avg", "to_equipment_avg",
              "other_avg", "transfers_avg"):
        ws.append(["monthly_%s" % k, data["monthly_overview"].get(k, 0)])
    ds = data.get("deliveries_summary") or {}
    for k in ("confirmed_count", "below_threshold_count", "below_threshold_pct",
              "band_minus1_to_0_count", "avg_variance_pct",
              "min_variance_pct", "max_variance_pct", "avg_daily_volume"):
        ws.append(["deliv_%s" % k, ds.get(k, 0)])

    ws = wb.create_sheet(SHEET_CONSOLIDATED)
    ws.append(["site", "opening", "deliveries", "transactions", "closing"])
    c = data["consolidated"]
    ws.append([c["site"], c["opening"], c["deliveries"], c["transactions"],
               c["closing"]])

    ws = wb.create_sheet(SHEET_TRUCKS)
    ws.append(["site", "deliveries", "transactions"])
    for t in data["service_trucks"]:
        ws.append([t["site"], t["deliveries"], t["transactions"]])

    ws = wb.create_sheet(SHEET_DELIVERIES)
    ws.append(["date", "tickets", "inlet"])
    for d in data["deliveries_daily"]:
        ws.append([d.get("date", ""), d.get("tickets", 0), d.get("inlet", 0)])

    ws = wb.create_sheet(SHEET_DISPENSING)
    ws.append(["field", "value"])
    ws.append(["to_equipment", data["dispensing"]["to_equipment"]])
    ws.append(["other", data["dispensing"]["other"]])
    ws.append(["transfers", data["dispensing"]["transfers"]])

    ws = wb.create_sheet(SHEET_WEEKLY)
    ws.append(["week", "date", "transactions", "to_equipment", "transfers"])
    for w in data["weekly_breakdown"]:
        ws.append([w.get("week"), w.get("date"), w.get("transactions"),
                   w.get("to_equipment"), w.get("transfers")])

    ws = wb.create_sheet(SHEET_TREND)
    ws.append(["date", "delivery_pct", "recon_pct"])
    for pt in data["monthly_trend"]:
        if len(pt) >= 3:
            ws.append([pt[0], pt[1], pt[2]])

    ws = wb.create_sheet(SHEET_HIST)
    ws.append(["label", "transactions", "to_equipment", "transfers", "other"])
    for row in data["historical_outflows"]:
        ws.append(list(row) + [0] * (5 - len(row)))

    ws = wb.create_sheet(SHEET_TEXTS)
    ws.append(["field", "value"])
    ws.append(["recon_sentence", data["recon_sentence"]])
    ws.append(["delivery_narrative", data["delivery_narrative"]])
    ws.append(["considerations", "\n".join(data["considerations"])])

    ws = wb.create_sheet(SHEET_TASKS)
    ws.append(["task"])
    for t in data["tasks_notes"]:
        ws.append([t])

    wb.save(path)


def load_excel(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    data = default_data()

    if SHEET_META in wb.sheetnames:
        meta = {row[0]: row[1] for row in wb[SHEET_META].iter_rows(
            min_row=2, values_only=True) if row and row[0]}
        for k in ("period_full", "month_label", "cover_date"):
            if meta.get(k):
                data[k] = str(meta[k])
        if "tank_log_min_value" in meta:
            data["tank_log_stats"]["min_value"] = _num(meta["tank_log_min_value"])
        if "tank_log_min_date" in meta:
            data["tank_log_stats"]["min_date"] = str(meta["tank_log_min_date"] or "")
        if "tank_log_average" in meta:
            data["tank_log_stats"]["average"] = _num(meta["tank_log_average"])
        for k in ("transactions_avg", "to_equipment_avg",
                  "other_avg", "transfers_avg"):
            mk = "monthly_%s" % k
            if mk in meta:
                data["monthly_overview"][k] = _num(meta[mk])
        for k in ("confirmed_count", "below_threshold_count",
                  "below_threshold_pct", "band_minus1_to_0_count",
                  "avg_variance_pct", "min_variance_pct",
                  "max_variance_pct", "avg_daily_volume"):
            mk = "deliv_%s" % k
            if mk in meta:
                data["deliveries_summary"][k] = _num(meta[mk])

    if SHEET_CONSOLIDATED in wb.sheetnames:
        for row in wb[SHEET_CONSOLIDATED].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                data["consolidated"] = {
                    "site": str(row[0]),
                    "opening": _num(row[1]),
                    "deliveries": _num(row[2]),
                    "transactions": _num(row[3]),
                    "closing": _num(row[4]),
                }
                break

    if SHEET_TRUCKS in wb.sheetnames:
        trucks = []
        for row in wb[SHEET_TRUCKS].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                trucks.append({
                    "site": str(row[0]),
                    "deliveries": _num(row[1]),
                    "transactions": _num(row[2]),
                })
        if trucks:
            data["service_trucks"] = trucks

    if SHEET_DELIVERIES in wb.sheetnames:
        delivs = []
        for row in wb[SHEET_DELIVERIES].iter_rows(min_row=2, values_only=True):
            if row and (row[0] or row[1] or row[2]):
                delivs.append({
                    "date": str(row[0] or ""),
                    "tickets": _num(row[1]),
                    "inlet": _num(row[2]),
                })
        data["deliveries_daily"] = delivs

    if SHEET_DISPENSING in wb.sheetnames:
        m = {row[0]: row[1] for row in wb[SHEET_DISPENSING].iter_rows(
            min_row=2, values_only=True) if row and row[0]}
        data["dispensing"] = {
            "to_equipment": _num(m.get("to_equipment")),
            "other": _num(m.get("other")),
            "transfers": _num(m.get("transfers")),
        }

    if SHEET_WEEKLY in wb.sheetnames:
        rows = []
        for r in wb[SHEET_WEEKLY].iter_rows(min_row=2, values_only=True):
            if r and r[0] is not None:
                rows.append({
                    "week": int(_num(r[0])),
                    "date": str(r[1] or ""),
                    "transactions": _num(r[2]),
                    "to_equipment": _num(r[3]),
                    "transfers": _num(r[4]),
                })
        data["weekly_breakdown"] = rows

    if SHEET_TREND in wb.sheetnames:
        pts = []
        for r in wb[SHEET_TREND].iter_rows(min_row=2, values_only=True):
            if r and r[0] is not None:
                pts.append((r[0], _num(r[1]), _num(r[2])))
        data["monthly_trend"] = pts

    if SHEET_HIST in wb.sheetnames:
        hist = []
        for r in wb[SHEET_HIST].iter_rows(min_row=2, values_only=True):
            if r and r[0] is not None:
                hist.append((str(r[0]), _num(r[1]), _num(r[2]),
                              _num(r[3]), _num(r[4])))
        data["historical_outflows"] = hist

    if SHEET_TEXTS in wb.sheetnames:
        m = {row[0]: row[1] for row in wb[SHEET_TEXTS].iter_rows(
            min_row=2, values_only=True) if row and row[0]}
        data["recon_sentence"] = str(m.get("recon_sentence") or "")
        data["delivery_narrative"] = str(m.get("delivery_narrative") or "")
        cons = str(m.get("considerations") or "")
        if cons.strip():
            data["considerations"] = [c.strip() for c in cons.split("\n")
                                      if c.strip()]

    if SHEET_TASKS in wb.sheetnames:
        tasks = []
        for row in wb[SHEET_TASKS].iter_rows(min_row=2, values_only=True):
            if row and row[0] and str(row[0]).strip():
                tasks.append(str(row[0]).strip())
        if tasks:
            data["tasks_notes"] = tasks

    return data


# --------------------------------------------------------------------------
# Calculos derivados
# --------------------------------------------------------------------------

def recon_pct_value(consolidated_raw: dict) -> float:
    """Variance % consolidado del mes, en escala 0-100."""
    opening = _num(consolidated_raw.get("opening"))
    deliveries = _num(consolidated_raw.get("deliveries"))
    transactions = _num(consolidated_raw.get("transactions"))
    closing = _num(consolidated_raw.get("closing"))
    if not transactions:
        return 0.0
    calc_stock = opening + deliveries - transactions
    return (closing - calc_stock) / transactions * 100.0


def delivery_pct_value(daily_rows: list) -> float:
    """(Sum tickets - Sum inlet) / Sum tickets * 100."""
    tot_inlet = sum(_num(r.get("inlet")) for r in daily_rows or [])
    tot_tick = sum(_num(r.get("tickets")) for r in daily_rows or [])
    if not tot_tick:
        return 0.0
    return (tot_inlet - tot_tick) / tot_tick * 100.0


def compute_consolidated_row(raw: dict) -> dict:
    opening = _num(raw["opening"])
    deliveries = _num(raw["deliveries"])
    transactions = _num(raw["transactions"])
    closing = _num(raw["closing"])
    calc_stock = opening + deliveries - transactions
    net_change = closing - opening
    variance = closing - calc_stock
    pct = (variance / transactions * 100.0) if transactions else 0.0
    return {
        "site": raw["site"],
        "opening": _fmt_dec2(opening),
        "deliveries": _fmt_dec2(deliveries),
        "transactions": fmt_int(transactions),
        "calc_stock": _fmt_dec2(calc_stock),
        "net_change": _fmt_dec2(net_change),
        "closing": _fmt_dec2(closing),
        "variance": _fmt_dec2(variance),
        "pct": fmt_pct(pct, 2),
        "_pct_value": pct,
    }


def _fmt_dec2(value) -> str:
    """123456.789 -> '123.456,79'  (miles con punto, decimal con coma)."""
    n = _num(value)
    s = "{:,.2f}".format(n)
    # cambia 1,234.56 -> 1.234,56
    return s.replace(",", "_").replace(".", ",").replace("_", ".")


def compute_deliveries_daily_rows(rows: list) -> tuple:
    """Devuelve (rows_fmt, totals)."""
    out, sum_inlet, sum_tick = [], 0.0, 0.0
    for r in rows:
        inlet = _num(r.get("inlet"))
        tick = _num(r.get("tickets"))
        var = inlet - tick
        pct = (var / tick * 100.0) if tick else 0.0
        sum_inlet += inlet
        sum_tick += tick
        out.append({
            "date": str(r.get("date", "")),
            "tickets": fmt_int(tick),
            "inlet": fmt_int(inlet),
            "variance": fmt_int(var),
            "pct": fmt_pct(pct, 2),
        })
    tot_var = sum_inlet - sum_tick
    tot_pct = (tot_var / sum_tick * 100.0) if sum_tick else 0.0
    return out, {
        "tickets": fmt_int(sum_tick),
        "inlet": fmt_int(sum_inlet),
        "variance": fmt_int(tot_var),
        "pct": fmt_pct(tot_pct, 2),
    }


def compute_truck_rows(rows: list) -> list:
    out = []
    for r in rows:
        out.append({
            "site": str(r["site"]),
            "deliveries": fmt_int(_num(r["deliveries"])),
            "transactions": fmt_int(_num(r["transactions"])),
        })
    return out


def compute_weekly_breakdown_rows(rows: list) -> list:
    out = []
    for r in rows:
        out.append({
            "week": str(r.get("week", "")),
            "date": str(r.get("date", "")),
            "transactions": fmt_int(_num(r.get("transactions"))),
            "to_equipment": fmt_int(_num(r.get("to_equipment"))),
            "transfers": fmt_int(_num(r.get("transfers"))),
        })
    return out


# --------------------------------------------------------------------------
# Construccion del contexto para docxtpl
# --------------------------------------------------------------------------

def _rt(parts) -> RichText:
    rt = RichText()
    for item in parts:
        if len(item) == 2:
            text, bold = item
            rt.add(str(text), bold=bool(bold), font=REPORT_FONT)
        else:
            text, bold, size = item
            rt.add(str(text), bold=bool(bold), font=REPORT_FONT, size=size)
    return rt


def _delivery_narrative_rt(summary: dict) -> RichText:
    """N confirmed deliveries. M deliveries below threshold -0,50%.
    The next sentence covers band -1..0%."""
    n = int(_num(summary.get("confirmed_count")))
    below = int(_num(summary.get("below_threshold_count")))
    band = int(_num(summary.get("band_minus1_to_0_count")))
    thr = _num(summary.get("below_threshold_pct"))
    thr_text = fmt_pct(thr, 2)

    deliv_word = "deliveries" if n != 1 else "delivery"
    rt = RichText()
    rt.add("%d confirmed %s. " % (n, deliv_word), bold=False, font=REPORT_FONT)
    rt.add("%d %s of these are below the %s threshold."
           % (below, deliv_word, thr_text), bold=True, font=REPORT_FONT)
    return rt


def _delivery_band_rt(summary: dict) -> RichText:
    band = int(_num(summary.get("band_minus1_to_0_count")))
    return _rt([
        ("%d deliveries were between " % band, False),
        ("-1,0%", True), (" and ", False), ("0%", True), (".", False),
    ])


def _delivery_stats_rt(summary: dict) -> RichText:
    avg = _num(summary.get("avg_variance_pct"))
    mn = _num(summary.get("min_variance_pct"))
    mx = _num(summary.get("max_variance_pct"))
    avg_vol = _num(summary.get("avg_daily_volume"))
    return _rt([
        ("The average variance is ", False),
        (fmt_pct(avg, 2), True),
        (", with minimum variances of ", False),
        (fmt_pct(mn, 2), True),
        (" and maximum variances of ", False),
        (fmt_pct(mx, 2), True),
        (". The average daily volume was ", False),
        (fmt_int(avg_vol), True), (" liters", True), (".", False),
    ])


def build_context(data: dict, image_paths: dict, tpl: DocxTemplate) -> dict:
    period = data["period_full"]
    month_label = data.get("month_label", "")
    ctx = {
        "period_full": period,
        "month_label": month_label,
        "cover_date": data.get("cover_date", ""),
    }

    cons = compute_consolidated_row(data["consolidated"])
    ctx["cons"] = cons

    rows, total = compute_deliveries_daily_rows(data["deliveries_daily"])
    ctx["deliveries"] = rows
    ctx["dtot"] = total

    ctx["trucks"] = compute_truck_rows(data["service_trucks"])
    ctx["weekly"] = compute_weekly_breakdown_rows(data["weekly_breakdown"])

    disp = data["dispensing"]
    eq = _num(disp["to_equipment"])
    ot = _num(disp["other"])
    tr = _num(disp["transfers"])
    total_disp = eq + ot + tr
    ctx["disp_to_equipment"] = fmt_int(eq)
    ctx["disp_other"] = fmt_int(ot)
    ctx["disp_transfers"] = fmt_int(tr)
    ctx["disp_total"] = fmt_int(total_disp)

    ctx["disp_total_line"] = _rt([
        ("A total ", False), ("Transactions", True), (" of ", False),
        (fmt_int(total_disp), True), (" L", True),
        (" were reported.", False),
    ])
    ctx["disp_equipment_line"] = _rt([
        ("Dispensing to Equipment: ", True),
        (fmt_int(eq), True), (" L", True), (".", False)])
    ctx["disp_other_line"] = _rt([
        ("Dispensing to other equipment ", True),
        (fmt_int(ot), True), (" L", True), (".", False)])
    ctx["disp_transfers_line"] = _rt([
        ("Transfers (Service Trucks) of ", True),
        (fmt_int(tr), True), (" L", True), (".", False)])

    stats = data["tank_log_stats"]
    ctx["tank_log_summary"] = _rt([
        ("Minimum Value in Summary Level Tank Diesel on ", False),
        (str(stats["min_date"]), True),
        (" ", False),
        (fmt_int(stats["min_value"]), True), (" L", True),
        (" and the average for the month ", False),
        (fmt_int(stats["average"]), True), (" L", True),
        (".", False),
    ])

    ovw = data["monthly_overview"]
    ctx["monthly_transactions_line"] = _rt([
        ("Transaction:", True),
        (" The average within the analyzed range is ", False),
        (fmt_int(ovw.get("transactions_avg", 0)), True), (" L", True),
        (".", False),
    ])
    ctx["monthly_to_equipment_line"] = _rt([
        ("To equipment:", True),
        (" The average within the analyzed range is ", False),
        (fmt_int(ovw.get("to_equipment_avg", 0)), True), (" L", True),
        (".", False),
    ])
    ctx["monthly_other_line"] = _rt([
        ("Distribution of other Dispenses:", True),
        (" The average within the analyzed range is ", False),
        (fmt_int(ovw.get("other_avg", 0)), True), (" L", True),
        (".", False),
    ])
    ctx["monthly_transfers_line"] = _rt([
        ("Distribution Transfers:", True),
        (" The average for this category is ", False),
        (fmt_int(ovw.get("transfers_avg", 0)), True), (" L", True),
        (".", False),
    ])

    # Recon sentence
    recon_custom = (data.get("recon_sentence") or "").strip()
    if recon_custom:
        ctx["recon_sentence"] = _rt([(recon_custom, False)])
    else:
        ctx["recon_sentence"] = _rt([
            ("The tank reconciliation for the period from ", False),
            (period, True),
            (" has a total variance of ", False),
            (cons["pct"], True),
            (". To have a better understanding of how reconciliation comes "
             "about, we will introduce you to some terms and instruments "
             "used in this report.", False),
        ])

    # Deliveries narrative auto si esta vacia
    narrative_custom = (data.get("delivery_narrative") or "").strip()
    summary = data.get("deliveries_summary") or {}
    if narrative_custom:
        ctx["delivery_narrative"] = _rt([(narrative_custom, False)])
    else:
        ctx["delivery_narrative"] = _delivery_narrative_rt(summary)
    ctx["delivery_band_line"] = _delivery_band_rt(summary)
    ctx["delivery_stats_line"] = _delivery_stats_rt(summary)

    ctx["considerations"] = list(data.get("considerations") or
                                  DEFAULT_CONSIDERATIONS)
    ctx["tasks_notes"] = list(data.get("tasks_notes") or [])

    # Figuras
    for fig in FIGURE_KEYS:
        path = (image_paths or {}).get(fig)
        if path and os.path.isfile(path):
            ctx[fig] = InlineImage(tpl, path, width=figure_width(fig))
        else:
            ctx[fig] = ""

    return ctx


def _build_auto_charts(data: dict, image_paths: dict, tmpdir: str) -> dict:
    """Genera las figuras que se construyen automaticamente con matplotlib.

    Las claves del dict resultante son las mismas que image_paths.
    Solo se genera la figura si NO viene una imagen custom en image_paths."""
    image_paths = image_paths or {}
    out = {}

    if not image_paths.get("fig1"):
        path = os.path.join(tmpdir, "fig1.png")
        charts.delivery_daily_bar_chart(data["deliveries_daily"], path)
        out["fig1"] = path

    if not image_paths.get("fig2"):
        path = os.path.join(tmpdir, "fig2.png")
        disp = data["dispensing"]
        charts.outflow_donut_chart(
            _num(disp["to_equipment"]), _num(disp["other"]),
            _num(disp["transfers"]), path,
            month_label=data.get("month_label", ""))
        out["fig2"] = path

    tank_points = data.get("tank_trend")
    if tank_points and not image_paths.get("fig3"):
        path = os.path.join(tmpdir, "fig3.png")
        charts.daily_tank_log_chart(
            tank_points, "Tank LFO - Main Tank", path,
            data.get("tank_safe_fill"))
        out["fig3"] = path

    trend = data.get("monthly_trend")
    if trend and not image_paths.get("fig4"):
        path = os.path.join(tmpdir, "fig4.png")
        charts.monthly_recon_trend_chart(
            trend, "Reconciliation for the product Diesel", path)
        out["fig4"] = path

    if trend and not image_paths.get("fig5"):
        path = os.path.join(tmpdir, "fig5.png")
        charts.delivery_vs_recon_chart(
            trend, "Delivery vs. Reconciliation for the product Diesel", path)
        out["fig5"] = path

    if data.get("weekly_breakdown") and not image_paths.get("fig6"):
        path = os.path.join(tmpdir, "fig6.png")
        charts.weekly_breakdown_bars(
            data["weekly_breakdown"],
            "Monthly Breakdown of Fuel Consumption: %s"
            % data.get("month_label", ""),
            path)
        out["fig6"] = path

    if data.get("historical_outflows") and not image_paths.get("fig7"):
        path = os.path.join(tmpdir, "fig7.png")
        charts.historical_outflows_bars(
            data["historical_outflows"],
            "Historical Transaction Volume by Type",
            path)
        out["fig7"] = path

    return out


_W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def sanitize_content_controls(docx_path: str) -> int:
    """Repara la causa del error de Word 'contenido no legible' heredada del
    reporte original.

    El banner del encabezado trae un control de contenido (`<w:sdt>`) de tipo
    *texto plano* enlazado a la propiedad Title (`<w:dataBinding>`) que, en vez
    de texto, envuelve una IMAGEN (`<w:drawing>` con el logo). Un control de
    texto plano no puede contener un dibujo: ese modelo de contenido invalido
    es lo que hace que Word pida "recuperar el contenido" al abrir el .docx.

    La correccion desenvuelve esos controles: elimina el `<w:sdt>`/`<w:sdtPr>`
    y deja su contenido interno (el parrafo con el logo) directamente en su
    lugar, conservando la apariencia del encabezado.

    Opera sobre document.xml y todos los headers/footers del paquete .docx.
    Devuelve cuantos controles se desenvolvieron. Idempotente."""
    from lxml import etree

    def qn(tag):
        return "{%s}%s" % (_W_NS, tag)

    with zipfile.ZipFile(docx_path, "r") as zin:
        infos = zin.infolist()
        contents = {i.filename: zin.read(i.filename) for i in infos}

    changed = 0
    for name in list(contents):
        if not (name.startswith("word/") and name.endswith(".xml")):
            continue
        if not ("header" in name or "footer" in name
                or name == "word/document.xml"):
            continue
        try:
            root = etree.fromstring(contents[name])
        except etree.XMLSyntaxError:
            continue
        local = 0
        for sdt in list(root.iter(qn("sdt"))):
            pr = sdt.find(qn("sdtPr"))
            if pr is None:
                continue
            is_text = pr.find(qn("text")) is not None
            is_bound = pr.find(qn("dataBinding")) is not None
            has_drawing = sdt.find(".//" + qn("drawing")) is not None
            if not (has_drawing and (is_text or is_bound)):
                continue
            content = sdt.find(qn("sdtContent"))
            parent = sdt.getparent()
            if parent is None:
                continue
            idx = list(parent).index(sdt)
            parent.remove(sdt)
            if content is not None:
                for j, child in enumerate(list(content)):
                    parent.insert(idx + j, child)
            local += 1
        if local:
            contents[name] = etree.tostring(
                root, xml_declaration=True, encoding="UTF-8",
                standalone=True)
            changed += local

    if changed:
        tmp = docx_path + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for i in infos:
                zi = zipfile.ZipInfo(i.filename, date_time=i.date_time)
                zi.compress_type = zipfile.ZIP_DEFLATED
                zi.external_attr = i.external_attr
                zout.writestr(zi, contents[i.filename])
        os.replace(tmp, docx_path)
    return changed


def generate_report(data: dict, image_paths: dict,
                     template_path: str, output_path: str) -> str:
    if not os.path.isfile(template_path):
        raise FileNotFoundError(
            "No se encontro la plantilla Word: %s" % template_path)
    tmpdir = tempfile.mkdtemp(prefix="mdiesel_charts_")
    try:
        auto = _build_auto_charts(data, image_paths, tmpdir)
        images = dict(image_paths or {})
        for fig_key, path in auto.items():
            images[fig_key] = path
        tpl = DocxTemplate(template_path)
        ctx = build_context(data, images, tpl)
        tpl.render(ctx)
        tpl.save(output_path)
        # Red de seguridad: aunque la plantilla venga limpia, garantizamos que
        # el .docx final no arrastre el control de contenido invalido del
        # encabezado (causa del aviso "contenido no legible" de Word).
        sanitize_content_controls(output_path)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
    return output_path
