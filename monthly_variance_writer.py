# -*- coding: utf-8 -*-
"""
Upsert de una fila en la hoja 'Monthly Variance' del Excel historico y
ajuste del rango del chart embebido.

La hoja tiene la siguiente estructura:
  - Fila 1: titulo "Weekly variance change" (heredado).
  - Fila 2: header (Date | Delivery % | Recon %).
  - Fila 3+: datos, una fila por mes (la fecha es el 1er dia del mes).

Los charts embebidos referencian rangos relativos; los deslizamos a las
ultimas N filas terminando en la mas reciente, igual que en el weekly
variance_writer.
"""
from __future__ import annotations

import datetime
import re

import openpyxl


SHEET_NAME = "Monthly Variance"
HEADER_ROW = 2
DATA_START = HEADER_ROW + 1

COL_DATE = 1
COL_DELIVERY = 2
COL_RECON = 3

DATE_FMT = "d/m/yy;@"
PCT_FMT = "0.00"

CHART_WINDOW = 24


def _check_excel_not_locked(excel_path: str) -> None:
    try:
        with open(excel_path, "a+b"):
            pass
    except PermissionError:
        raise PermissionError(
            "El archivo Excel parece estar abierto en Microsoft Excel:\n"
            "  %s\n\nCierre el archivo en Excel y vuelva a intentar."
            % excel_path)


def _find_row_for_date(ws, day: datetime.date) -> int | None:
    for r in range(DATA_START, ws.max_row + 1):
        v = ws.cell(row=r, column=COL_DATE).value
        if isinstance(v, datetime.datetime) and v.date() == day:
            return r
        if isinstance(v, datetime.date) and v == day:
            return r
    return None


def _last_data_row(ws) -> int:
    last = HEADER_ROW
    for r in range(DATA_START, ws.max_row + 1):
        if ws.cell(row=r, column=COL_DATE).value is not None:
            last = r
    return last


_RANGE_RE = re.compile(r"\$([A-Z]+)\$\d+:\$([A-Z]+)\$\d+")


def _replace_rows(ref: str, start: int, end: int) -> str:
    return _RANGE_RE.sub(
        lambda mo: "$%s$%d:$%s$%d" % (mo.group(1), start, mo.group(2), end),
        ref)


def _slide_chart_window(ws, last_row: int, window: int = CHART_WINDOW) -> bool:
    charts = getattr(ws, "_charts", [])
    if not charts:
        return False
    start = max(DATA_START, last_row - window + 1)
    changed = False
    for ch in charts:
        for s in ch.series:
            for ref_holder in (s.val, s.cat):
                if ref_holder is None:
                    continue
                num = getattr(ref_holder, "numRef", None)
                if num is not None and num.f:
                    new_f = _replace_rows(num.f, start, last_row)
                    if new_f != num.f:
                        num.f = new_f
                        changed = True
                strref = getattr(ref_holder, "strRef", None)
                if strref is not None and strref.f:
                    new_f = _replace_rows(strref.f, start, last_row)
                    if new_f != strref.f:
                        strref.f = new_f
                        changed = True
    return changed


def upsert_month(excel_path: str, month_first: datetime.date,
                  delivery_pct: float, recon_pct: float) -> dict:
    """Inserta o actualiza la fila correspondiente a `month_first` y desliza
    la ventana del chart embebido."""
    _check_excel_not_locked(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return {"row": None, "action": "skipped_no_sheet"}
    ws = wb[SHEET_NAME]

    existing = _find_row_for_date(ws, month_first)
    if existing is not None:
        target = existing
        action = "updated"
    else:
        target = _last_data_row(ws) + 1
        action = "inserted"

    date_cell = ws.cell(row=target, column=COL_DATE,
                        value=datetime.datetime(month_first.year,
                                                 month_first.month,
                                                 month_first.day))
    date_cell.number_format = DATE_FMT

    deliv_cell = ws.cell(row=target, column=COL_DELIVERY,
                         value=round(float(delivery_pct), 2))
    deliv_cell.number_format = PCT_FMT

    recon_cell = ws.cell(row=target, column=COL_RECON,
                         value=round(float(recon_pct), 2))
    recon_cell.number_format = PCT_FMT

    last_row = _last_data_row(ws)
    window_start = max(DATA_START, last_row - CHART_WINDOW + 1)
    _slide_chart_window(ws, last_row, CHART_WINDOW)

    wb.save(excel_path)
    wb.close()
    return {"row": target, "action": action,
            "window": (window_start, last_row)}
