# -*- coding: utf-8 -*-
"""
Importa el CSV 'delivery_transaction_*.csv' (export de Veridapt) y lo
vuelca en la hoja 'delivery_transaction_*' del Excel historico
'Reconciliation Monthly.xlsx'.

Misma logica que diesel_report.delivery_csv_import; lo unico que cambia es
que el sheet destino del monthly se llama 'delivery_transaction_213_202312'
(arbitrario por su origen). La deteccion se hace por prefijo, asi que
ambos archivos funcionan igual.
"""
from __future__ import annotations

import csv
import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


SHEET_PREFIX = "delivery_transaction"


_HEADER_ALIASES = {
    "product":        ["product"],
    "collected":      ["collected at", "collected_at", "collected", "date"],
    "tank":           ["tank"],
    "docket":         ["docket number", "docket_number", "docket"],
    "supplier":       ["supplier"],
    "confirmed":      ["confirmed"],
    "type":           ["type"],
    "unit":           ["volume unit", "volume_unit", "unit"],
    "volume":         ["volume"],
    "docket_volume":  ["docket volume", "docket_volume"],
    "variance":       ["variance"],
    "variance_pct":   ["delivery variance %", "delivery variance",
                       "variance %", "%", "variance_pct"],
}

_COLLECTED_DATE_FMT = "mm-dd-yy"
_DOCKET_VOL_FMT = "General"
_VARIANCE_FMT = "#,##0.00"
_VARIANCE_PCT_FMT = "0.00%"


def _normalize_header(text) -> str:
    return str(text or "").strip().lower()


def _build_col_map(header_row: list) -> dict:
    out = {}
    for idx, cell in enumerate(header_row, start=1):
        name = _normalize_header(cell)
        if not name:
            continue
        for logical, aliases in _HEADER_ALIASES.items():
            if name in aliases and logical not in out:
                out[logical] = idx
                break
    return out


def _parse_csv_date(text: str):
    text = str(text or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _parse_csv_number(text: str) -> float:
    text = str(text or "").strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _read_csv_rows(csv_path: str) -> list:
    out = []
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            confirmed = str(row.get("Confirmed", "")).strip().lower()
            if confirmed != "yes":
                continue
            docket = str(row.get("Docket Number", "")).strip()
            if not docket:
                continue
            out.append({
                "product":  row.get("Product", ""),
                "collected": _parse_csv_date(row.get("Collected At", "")),
                "tank":     row.get("Tank", ""),
                "docket":   docket,
                "supplier": row.get("Supplier", ""),
                "confirmed": "Yes",
                "type":     row.get("Type", ""),
                "unit":     row.get("Volume Unit", ""),
                "volume":   _parse_csv_number(row.get("Volume", "")),
            })
    return out


def _find_sheet(wb) -> str | None:
    for name in wb.sheetnames:
        if name.lower().startswith(SHEET_PREFIX):
            return name
    return None


def _existing_dockets(ws, col_docket: int) -> dict:
    out = {}
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=col_docket).value
        if val is None:
            continue
        key = str(val).strip()
        if key:
            out[key] = row_idx
    return out


def _last_data_row(ws, col_anchor: int) -> int:
    last = 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=col_anchor).value not in (None, ""):
            last = row_idx
    return last


def _extend_tables_to_row(ws, target_row: int) -> None:
    try:
        tables = list(ws.tables.values())
    except AttributeError:
        return
    for table in tables:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        except Exception:
            continue
        if target_row > max_row:
            table.ref = "%s%d:%s%d" % (
                get_column_letter(min_col), min_row,
                get_column_letter(max_col), target_row)


def _check_excel_not_locked(excel_path: str) -> None:
    try:
        with open(excel_path, "a+b"):
            pass
    except PermissionError:
        raise PermissionError(
            "El archivo Excel parece estar abierto en Microsoft Excel:\n"
            "  %s\n\n"
            "Cierre el archivo en Excel y vuelva a intentar la carga."
            % excel_path)


def import_csv_to_excel(excel_path: str, csv_path: str) -> dict:
    """Importa el CSV y lo escribe en la hoja delivery_transaction_*.
    Devuelve estadisticas del upsert."""
    _check_excel_not_locked(excel_path)
    csv_rows = _read_csv_rows(csv_path)
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = _find_sheet(wb)
    if sheet_name is None:
        wb.close()
        return {"sheet": None, "inserted": 0, "updated": 0,
                "total_csv": len(csv_rows)}

    ws = wb[sheet_name]
    header_row = [ws.cell(row=1, column=c).value
                  for c in range(1, (ws.max_column or 1) + 1)]
    col_map = _build_col_map(header_row)

    if "docket" not in col_map or "volume" not in col_map:
        default_header = ["Product", "Collected At", "Tank", "Docket Number",
                          "Supplier", "Confirmed", "Type", "Volume Unit",
                          "Volume"]
        for c, txt in enumerate(default_header, start=1):
            ws.cell(row=1, column=c, value=txt)
        col_map = _build_col_map(default_header)

    col_docket = col_map["docket"]
    existing = _existing_dockets(ws, col_docket)
    last_row = _last_data_row(ws, col_docket)

    inserted = updated = 0
    for r in csv_rows:
        target = existing.get(r["docket"])
        if target is None:
            last_row += 1
            target = last_row
            inserted += 1
        else:
            updated += 1

        def _set(logical, value, number_format=None):
            col = col_map.get(logical)
            if col:
                cell = ws.cell(row=target, column=col, value=value)
                if number_format is not None:
                    cell.number_format = number_format

        _set("product", r["product"])
        _set("collected", r["collected"], number_format=_COLLECTED_DATE_FMT)
        _set("tank", r["tank"])
        _set("docket", r["docket"])
        _set("supplier", r["supplier"])
        _set("confirmed", r["confirmed"])
        _set("type", r["type"])
        _set("unit", r["unit"])
        _set("volume", r["volume"])

        col_vol = col_map.get("volume")
        if col_vol:
            vol_ref = "%s%d" % (get_column_letter(col_vol), target)
            _set("docket_volume",
                 "=IF(%s>35000,40000,34000)" % vol_ref,
                 number_format=_DOCKET_VOL_FMT)
            _set("variance",
                 "=Table1[[#This Row],[Docket Volume]]-%s" % vol_ref,
                 number_format=_VARIANCE_FMT)
            _set("variance_pct",
                 "=Table1[[#This Row],[Variance]]/"
                 "Table1[[#This Row],[Docket Volume]]",
                 number_format=_VARIANCE_PCT_FMT)

    _extend_tables_to_row(ws, last_row)
    wb.save(excel_path)
    wb.close()
    return {"sheet": sheet_name, "inserted": inserted, "updated": updated,
            "total_csv": len(csv_rows)}
