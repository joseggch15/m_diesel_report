# -*- coding: utf-8 -*-
"""
Lectura del Excel historico "Reconciliation Monthly.xlsx".

Hojas que conoce:
  - Recon_LFO              : una fila POR MES (date = 1er dia del mes) para el
                             Main Tank consolidado.
  - Recon_FuelTank 0846/47/48: igual, por service truck.
  - Recon_LFO Weekly       : reconciliacion semanal historica (sirve para el
                             "weekly breakdown" del mes y para el monthly
                             overview avgs).
  - Monthly Variance       : Date | Delivery % | Recon % (puntos para las
                             figuras 4 y 5).
  - OUTFLOWS / OT Weekly   : weekly Transactions/ToEq/Other/Transfers.
  - PVT OUTFLOW            : agregado historico mes-a-mes (Fig 7).
  - delivery_transaction_* : tickets historicos (Confirmed=Yes), una fila por
                             ticket. Se filtra por el mes objetivo.

Estructura de columnas en Recon_LFO / Recon_FuelTank (header en R1, sub-
header en R2, datos desde R3):
  A=Date, B=Site, C=Opening, D=Deliveries, E=Transactions, F=Calculated,
  G=Net change, H=Closing, I=Variance, J=%, L=To equipment, M=Other,
  N=Transfers.

Estructura de Recon_LFO Weekly: igual, una fila por semana cerrada.
"""
from __future__ import annotations

import calendar
import datetime

import openpyxl

import report_model as m

# Indices (0-based en iter_rows values_only).
COL_DATE = 0
COL_SITE = 1
COL_OPENING = 2
COL_DELIVERIES = 3
COL_TRANSACTIONS = 4
COL_CLOSING = 7
COL_TO_EQUIPMENT = 11
COL_OTHER = 12
COL_TRANSFERS = 13

# Hoja Monthly Variance.
MV_COL_DATE = 0
MV_COL_DELIVERY = 1
MV_COL_RECON = 2

# delivery_transaction (con cols calculadas docket_volume, variance, %).
DT_PRODUCT = 0
DT_COLLECTED = 1
DT_CONFIRMED = 5
DT_VOLUME = 7
DT_DOCKET_VOL = 8     # Docket Volume (40k o 34k)
DT_VARIANCE = 9       # docket - volume
DT_PCT = 10           # variance / docket (fraccion, NO %)

TRUCK_SHEETS = {
    "TFL0846": "Recon_FuelTank 0846",
    "TFL0847": "Recon_FuelTank 0847",
    "TFL0848": "Recon_FuelTank 0848",
}


class MonthRecord:
    __slots__ = ("opening", "deliveries", "transactions", "closing",
                 "to_equipment", "other", "transfers")

    def __init__(self):
        self.opening = 0.0
        self.deliveries = 0.0
        self.transactions = 0.0
        self.closing = 0.0
        self.to_equipment = 0.0
        self.other = 0.0
        self.transfers = 0.0


def _is_date(value):
    return isinstance(value, (datetime.datetime, datetime.date))


def _to_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


def _previous_month_first(d: datetime.date) -> datetime.date:
    """2026-05-01 -> 2026-04-01 (Veridapt convention: el row dated N+1 / 1
    representa el cierre del mes N)."""
    if d.month == 1:
        return d.replace(year=d.year - 1, month=12, day=1)
    return d.replace(month=d.month - 1, day=1)


def _read_recon_sheet(ws) -> dict:
    """Indexa una hoja Recon_* por fecha (1er dia del mes REPORTADO). El
    Excel guarda la fecha del row como 1er dia del mes SIGUIENTE (convencion
    Veridapt: el row dated 2026-05-01 contiene los totales de Abril). Aqui
    aplicamos -1 mes para que la clave sea el mes de reporte."""
    by_date = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        raw_day = _to_date(row[COL_DATE])
        if raw_day is None or raw_day.day != 1:
            continue
        # Filas sin datos (solo fecha) se saltan
        if row[COL_OPENING] is None and row[COL_CLOSING] is None:
            continue
        day = _previous_month_first(raw_day)
        rec = by_date.setdefault(day, MonthRecord())
        rec.opening += m._num(row[COL_OPENING])
        rec.deliveries += m._num(row[COL_DELIVERIES])
        rec.transactions += m._num(row[COL_TRANSACTIONS])
        rec.closing += m._num(row[COL_CLOSING])
        if len(row) > COL_TRANSFERS:
            rec.to_equipment += m._num(row[COL_TO_EQUIPMENT])
            rec.other += m._num(row[COL_OTHER])
            rec.transfers += m._num(row[COL_TRANSFERS])
    return by_date


def _read_weekly_sheet(ws) -> list:
    """Lee todas las filas semanales (Recon_LFO Weekly). Devuelve lista de
    (date, MonthRecord)."""
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        day = _to_date(row[COL_DATE])
        if day is None:
            continue
        if row[COL_OPENING] is None and row[COL_CLOSING] is None:
            continue
        rec = MonthRecord()
        rec.opening = m._num(row[COL_OPENING])
        rec.deliveries = m._num(row[COL_DELIVERIES])
        rec.transactions = m._num(row[COL_TRANSACTIONS])
        rec.closing = m._num(row[COL_CLOSING])
        if len(row) > COL_TRANSFERS:
            rec.to_equipment = m._num(row[COL_TO_EQUIPMENT])
            rec.other = m._num(row[COL_OTHER])
            rec.transfers = m._num(row[COL_TRANSFERS])
        rows.append((day, rec))
    rows.sort(key=lambda x: x[0])
    return rows


def _read_pvt_outflow(ws) -> list:
    """Lee la hoja 'PVT OUTFLOW'.

    Estructura:
      R1, R2 vacias
      R3: Row Labels | Sum of Transactions | To equipment | Transfers | Other
      R4+: filas con 'YYYY' (totales anuales) y bajo cada ano los meses
           abreviados (ene/feb/mar/...) o 'jan/26 feb/26 ...'.

    Solo nos interesan las filas mensuales (no las anuales). Devuelve lista
    de (label, transactions, to_equipment, transfers, other) en orden de
    aparicion. Filtra las filas cuyo label tiene 4 digitos (anos)."""
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        # Salta filas totalizadoras (solo digitos = ano).
        if label.replace(" ", "").isdigit() and len(label) == 4:
            continue
        # Salta "Grand Total"
        if label.lower().startswith("grand total"):
            continue
        rows.append((
            label,
            m._num(row[1]),  # Transactions
            m._num(row[2]),  # To equipment
            m._num(row[3]),  # Transfers
            m._num(row[4]) if len(row) > 4 else 0.0,  # Other
        ))
    return rows


def _read_monthly_variance(ws) -> list:
    """Lee la hoja 'Monthly Variance'. Aplica la misma convencion Veridapt:
    el row dated N+1/1 contiene los % del mes N. Devuelve [(month_first,
    delivery_pct, recon_pct), ...] en orden cronologico."""
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        raw_day = _to_date(row[MV_COL_DATE])
        if raw_day is None or raw_day.day != 1:
            continue
        day = _previous_month_first(raw_day)
        deliv = m._num(row[MV_COL_DELIVERY]) if len(row) > MV_COL_DELIVERY \
            else 0.0
        recon = m._num(row[MV_COL_RECON]) if len(row) > MV_COL_RECON else 0.0
        out.append((day, deliv, recon))
    out.sort(key=lambda x: x[0])
    return out


class MonthlyHistory:
    """Repositorio de lectura del Excel historico."""

    def __init__(self):
        self._lfo: dict = {}              # {date_first_of_month: MonthRecord}
        self._trucks: dict = {}            # {site: {date: MonthRecord}}
        self._lfo_weekly: list = []        # [(date, MonthRecord)]
        self._monthly_variance: list = []  # [(date, deliv_pct, recon_pct)]
        self._pvt_outflow: list = []       # [(label, tx, eq, tr, other)]
        self._deliveries: list = []        # tickets {dt, volume, docket, pct}
        self.source_path = None

    def load(self, path: str) -> None:
        wb = openpyxl.load_workbook(path, data_only=True)
        names = wb.sheetnames
        self.source_path = path

        if "Recon_LFO" in names:
            self._lfo = _read_recon_sheet(wb["Recon_LFO"])
        else:
            self._lfo = {}

        self._trucks = {}
        for site, sheet in TRUCK_SHEETS.items():
            if sheet in names:
                self._trucks[site] = _read_recon_sheet(wb[sheet])
            else:
                self._trucks[site] = {}

        self._lfo_weekly = []
        if "Recon_LFO Weekly" in names:
            self._lfo_weekly = _read_weekly_sheet(wb["Recon_LFO Weekly"])

        self._monthly_variance = []
        if "Monthly Variance" in names:
            self._monthly_variance = _read_monthly_variance(
                wb["Monthly Variance"])

        self._pvt_outflow = []
        if "PVT OUTFLOW" in names:
            self._pvt_outflow = _read_pvt_outflow(wb["PVT OUTFLOW"])

        # Tickets historicos (delivery_transaction_*).
        self._deliveries = []
        dt_sheet = next((s for s in names
                         if s.lower().startswith("delivery_transaction")),
                        None)
        if dt_sheet is not None:
            self._read_deliveries(wb[dt_sheet])

    def _read_deliveries(self, ws) -> None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            collected = row[DT_COLLECTED] if len(row) > DT_COLLECTED else None
            if not _is_date(collected):
                continue
            if len(row) <= DT_CONFIRMED or \
                    str(row[DT_CONFIRMED]).strip().lower() != "yes":
                continue
            volume = m._num(row[DT_VOLUME]) if len(row) > DT_VOLUME else 0.0
            docket = m._num(row[DT_DOCKET_VOL]) \
                if len(row) > DT_DOCKET_VOL and row[DT_DOCKET_VOL] is not None \
                else (40000.0 if volume > 35000 else 34000.0)
            pct_raw = m._num(row[DT_PCT]) if len(row) > DT_PCT \
                and row[DT_PCT] is not None else 0.0
            # Si el campo % esta como fraccion (-0.0046), normalizamos a %.
            # Si ya viene en %, lo dejamos.
            pct = pct_raw * 100.0 if -1.0 < pct_raw < 1.0 else pct_raw
            self._deliveries.append({
                "dt": collected if isinstance(collected, datetime.datetime)
                    else datetime.datetime.combine(collected, datetime.time()),
                "volume": volume,
                "docket": docket,
                "variance_pct": pct,
            })

    # ---------------- Queries -----------------

    def is_loaded(self) -> bool:
        return bool(self._lfo)

    def available_months(self) -> list:
        """Fechas (1er dia del mes) presentes en Recon_LFO, mas reciente
        primero."""
        return sorted(self._lfo.keys(), reverse=True)

    def lookup_lfo(self, month_first) -> MonthRecord | None:
        return self._lfo.get(month_first)

    def lookup_truck(self, site: str, month_first) -> MonthRecord | None:
        sheet = self._trucks.get(site, {})
        return sheet.get(month_first)

    def monthly_trend_until(self, month_first, limit: int = 24) -> list:
        """Puntos (date, deliv_pct, recon_pct) hasta el mes inclusive,
        ultimos `limit` meses."""
        pts = [p for p in self._monthly_variance if p[0] <= month_first]
        return pts[-limit:]

    def historical_outflows(self, limit: int = 24) -> list:
        """Ultimas `limit` filas mensuales del PVT OUTFLOW."""
        return self._pvt_outflow[-limit:]

    def deliveries_for_month(self, month_first) -> list:
        """Lista de tickets confirmados dentro del mes (datetime, vol, docket,
        pct)."""
        first, last = m.first_last_of_month(month_first)
        out = []
        for tk in self._deliveries:
            d = tk["dt"].date()
            if first <= d <= last:
                out.append(tk)
        return out

    def deliveries_daily_for_month(self, month_first) -> list:
        """Agrupa por dia los tickets del mes y devuelve lista con
        {date: 'd-mmm', tickets: sum(docket), inlet: sum(volume)}."""
        tickets = self.deliveries_for_month(month_first)
        per_day = {}
        for tk in tickets:
            d = tk["dt"].date()
            acc = per_day.setdefault(d, [0.0, 0.0])
            acc[0] += tk["docket"]
            acc[1] += tk["volume"]
        out = []
        for d in sorted(per_day):
            sum_doc, sum_vol = per_day[d]
            out.append({
                "_date": d,
                "date": "%d-%s" % (d.day, d.strftime("%b").lower()),
                "tickets": round(sum_doc, 2),
                "inlet": round(sum_vol, 2),
            })
        return out

    def deliveries_summary_for_month(self, month_first,
                                      threshold_pct: float = -0.50) -> dict:
        """Conteos y promedios para la narrativa.
          - confirmed_count: total tickets
          - below_threshold_count: tickets con pct < threshold_pct
          - band_minus1_to_0_count: tickets con -1% < pct <= 0%
          - avg_variance_pct, min_variance_pct, max_variance_pct
          - avg_daily_volume: sum(volume) / dias_del_mes
        """
        tickets = self.deliveries_for_month(month_first)
        n = len(tickets)
        pcts = [t["variance_pct"] for t in tickets]
        sum_vol = sum(t["volume"] for t in tickets)
        first, last = m.first_last_of_month(month_first)
        days = (last - first).days + 1
        # "Below the threshold" en el manual significa MEJOR que el umbral,
        # es decir, mas cerca de cero (menor en magnitud). Para threshold
        # -0,50%: cuentan los tickets con pct > -0,50%.
        below = sum(1 for p in pcts if p > threshold_pct)
        band = sum(1 for p in pcts if -1.0 < p <= 0.0)
        if pcts:
            avg = sum(pcts) / len(pcts)
            mn = min(pcts)
            mx = max(pcts)
        else:
            avg = mn = mx = 0.0
        return {
            "confirmed_count": n,
            "below_threshold_count": below,
            "below_threshold_pct": threshold_pct,
            "band_minus1_to_0_count": band,
            "avg_variance_pct": avg,
            "min_variance_pct": mn,
            "max_variance_pct": mx,
            "avg_daily_volume": sum_vol / days if days else 0.0,
        }

    def weekly_breakdown_for_month(self, month_first) -> list:
        """Devuelve la lista de filas {week, date 'dd/mm/yy', transactions,
        to_equipment, transfers} correspondiente al mes.

        Estrategia: en el reporte manual cada fila representa el cierre de
        una semana cuyo periodo cae dentro del mes (incluso si el cierre cae
        el primer dia del siguiente mes, e.g., '06/04/26' es cierre Lun
        despues de la primera semana). Tomamos las filas de Recon_LFO Weekly
        que tengan fecha en [first, last_of_next_week]."""
        first, last = m.first_last_of_month(month_first)
        # El reporte de Abril usa 06/04, 13/04, 20/04, 27/04, 30/04. Todas las
        # filas weekly con fecha en [first, last] entran (incluida una al
        # cierre del mes si la fecha es last day).
        out = []
        week_idx = 0
        for d, rec in self._lfo_weekly:
            if first <= d <= last:
                week_idx += 1
                out.append({
                    "week": week_idx,
                    "date": d.strftime("%d/%m/%y"),
                    "_date": d,
                    "transactions": round(rec.transactions, 2),
                    "to_equipment": round(rec.to_equipment, 2),
                    "transfers": round(rec.transfers, 2),
                })
        return out

    def monthly_overview_from_breakdown(self, weekly_rows: list) -> dict:
        """Calcula los avgs del bloque 'Monthly Transactions overview'
        a partir de las filas semanales del mes."""
        n = len(weekly_rows)
        if n == 0:
            return {"transactions_avg": 0, "to_equipment_avg": 0,
                    "other_avg": 0, "transfers_avg": 0}
        return {
            "transactions_avg":
                round(sum(r["transactions"] for r in weekly_rows) / n, 0),
            "to_equipment_avg":
                round(sum(r["to_equipment"] for r in weekly_rows) / n, 0),
            "other_avg": 0,  # En el manual siempre 0
            "transfers_avg":
                round(sum(r["transfers"] for r in weekly_rows) / n, 0),
        }


def apply_month_to_data(data: dict, store: MonthlyHistory,
                         month_first: datetime.date,
                         threshold_pct: float = -0.50) -> dict:
    """Vuelca los datos del mes `month_first` sobre `data`. Devuelve un
    resumen indicando que se encontro y que no."""
    summary = {"lfo": False, "trucks": [], "trucks_missing": [],
               "deliveries_count": 0, "weekly_rows": 0,
               "trend_points": 0, "historical_points": 0}

    rec = store.lookup_lfo(month_first)
    if rec is not None:
        summary["lfo"] = True
        data["consolidated"] = {
            "site": m.SITE_KEY,
            "opening": round(rec.opening, 2),
            "deliveries": round(rec.deliveries, 2),
            "transactions": round(rec.transactions, 2),
            "closing": round(rec.closing, 2),
        }
        data["dispensing"] = {
            "to_equipment": round(rec.to_equipment, 2),
            "other": round(rec.other, 2),
            "transfers": round(rec.transfers, 2),
        }

    trucks = []
    for site in m.SERVICE_TRUCKS:
        tr = store.lookup_truck(site, month_first)
        if tr is None:
            summary["trucks_missing"].append(site)
            continue
        summary["trucks"].append(site)
        trucks.append({
            "site": site,
            "deliveries": round(tr.deliveries, 2),
            "transactions": round(tr.transactions, 2),
        })
    if trucks:
        data["service_trucks"] = trucks

    # Deliveries diarias del mes.
    daily = store.deliveries_daily_for_month(month_first)
    data["deliveries_daily"] = daily
    summary["deliveries_count"] = sum(1 for tk in store.deliveries_for_month(
        month_first))

    # Narrativa de deliveries.
    data["deliveries_summary"] = store.deliveries_summary_for_month(
        month_first, threshold_pct)

    # Weekly breakdown.
    weekly = store.weekly_breakdown_for_month(month_first)
    data["weekly_breakdown"] = weekly
    summary["weekly_rows"] = len(weekly)
    data["monthly_overview"] = store.monthly_overview_from_breakdown(weekly)

    # Monthly trend (Fig 4 y 5).
    trend = store.monthly_trend_until(month_first, limit=24)
    data["monthly_trend"] = trend
    summary["trend_points"] = len(trend)

    # Historical outflows (Fig 7).
    hist = store.historical_outflows(limit=24)
    data["historical_outflows"] = hist
    summary["historical_points"] = len(hist)

    # Period + cover.
    first, last = m.first_last_of_month(month_first)
    data["period_full"] = m.format_period_full(first, last)
    data["month_label"] = m.format_month_label(first)
    data["cover_date"] = m.format_cover_date(last)

    return summary
