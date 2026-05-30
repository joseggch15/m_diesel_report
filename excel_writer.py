# -*- coding: utf-8 -*-
"""
Escribe los datos extraidos del PDF Veridapt en el Excel historico
'Reconciliation Monthly.xlsx', hojas Recon_LFO / Recon_FuelTank.

Para la fecha indicada (1er dia del mes correspondiente) se inserta una nueva
fila o se actualiza la existente, respetando estilos del template del row
anterior.

Las columnas E, F, G, I, J se escriben como NUMEROS (no formulas) para que al
re-leer el archivo con data_only=True no devuelvan None.
"""
from __future__ import annotations

import copy
import datetime

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


SITE_TO_SHEET = {
    "LFO":     ("Recon_LFO",            "LFO"),
    "TFL0846": ("Recon_FuelTank 0846",  "TFL0846"),
    "TFL0847": ("Recon_FuelTank 0847",  "TFL0847"),
    "TFL0848": ("Recon_FuelTank 0848",  "TFL0848"),
}

# Las hojas tienen 2 filas de header (R1 y R2). Datos desde R3.
HEADER_ROWS = 2
DATA_START = HEADER_ROWS + 1

_COL_DATE = 1          # A
_COL_SITE = 2          # B
_COL_OPENING = 3       # C
_COL_DELIVERIES = 4    # D
_COL_TRANSACTIONS = 5  # E
_COL_CALC_STOCK = 6    # F
_COL_NET_CHANGE = 7    # G
_COL_CLOSING = 8       # H
_COL_VARIANCE = 9      # I
_COL_PCT = 10          # J
_COL_TO_EQUIPMENT = 12  # L
_COL_OTHER = 13         # M
_COL_TRANSFERS = 14     # N


def _last_data_row(ws) -> int:
    last = HEADER_ROWS
    for row_idx in range(DATA_START, ws.max_row + 1):
        if isinstance(ws.cell(row=row_idx, column=_COL_DATE).value,
                      (datetime.datetime, datetime.date)):
            last = row_idx
    return last


def _find_row_by_date(ws, target_date: datetime.date) -> int | None:
    for row_idx in range(DATA_START, ws.max_row + 1):
        cell_date = ws.cell(row=row_idx, column=_COL_DATE).value
        if isinstance(cell_date, datetime.datetime):
            if cell_date.date() == target_date:
                return row_idx
        elif isinstance(cell_date, datetime.date):
            if cell_date == target_date:
                return row_idx
    return None


def _copy_row_template(ws, source_row: int, target_row: int,
                       max_col: int = 16) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        val = src.value
        if isinstance(val, str) and val.startswith("="):
            col_letter = get_column_letter(col)
            try:
                tr = Translator(val, origin="%s%d" % (col_letter, source_row))
                tgt.value = tr.translate_formula(
                    "%s%d" % (col_letter, target_row))
            except Exception:
                tgt.value = val
        if src.has_style:
            tgt.font = copy.copy(src.font)
            tgt.fill = copy.copy(src.fill)
            tgt.border = copy.copy(src.border)
            tgt.alignment = copy.copy(src.alignment)
            tgt.number_format = src.number_format
            tgt.protection = copy.copy(src.protection)


def _compute_derived_lfo(info: dict) -> dict:
    """LFO Main + Virtual: Inflow/Outflow del Product Summary del PDF
    (Veridapt ya consolida sub-tanques)."""
    opening = float(info.get("opening", 0) or 0)
    inflow = float(info.get("inflow", 0) or 0)
    outflow = float(info.get("outflow", 0) or 0)
    closing = float(info.get("closing", 0) or 0)
    transactions = outflow
    calc_stock = opening + inflow - transactions
    net_change = closing - opening
    variance = closing - calc_stock
    pct = (variance / transactions) if transactions else 0.0
    return {
        "transactions": transactions,
        "calc_stock": calc_stock,
        "net_change": net_change,
        "variance": variance,
        "pct": pct,
    }


def _write_one_lfo(ws, target_row: int, month_first: datetime.date,
                   site_value: str, info: dict) -> None:
    d = _compute_derived_lfo(info)
    ws.cell(row=target_row, column=_COL_DATE,
            value=datetime.datetime.combine(month_first, datetime.time()))
    ws.cell(row=target_row, column=_COL_SITE, value=site_value)
    ws.cell(row=target_row, column=_COL_OPENING, value=info["opening"])
    ws.cell(row=target_row, column=_COL_DELIVERIES, value=info["inflow"])
    ws.cell(row=target_row, column=_COL_TRANSACTIONS, value=d["transactions"])
    ws.cell(row=target_row, column=_COL_CALC_STOCK, value=d["calc_stock"])
    ws.cell(row=target_row, column=_COL_NET_CHANGE, value=d["net_change"])
    ws.cell(row=target_row, column=_COL_CLOSING, value=info["closing"])
    ws.cell(row=target_row, column=_COL_VARIANCE, value=d["variance"])
    ws.cell(row=target_row, column=_COL_PCT, value=d["pct"])
    ws.cell(row=target_row, column=_COL_TO_EQUIPMENT,
            value=info["to_equipment"])
    ws.cell(row=target_row, column=_COL_OTHER, value=info["other_dispenses"])
    ws.cell(row=target_row, column=_COL_TRANSFERS, value=info["transfers_out"])


def _write_one_truck(ws, target_row: int, month_first: datetime.date,
                     site_value: str, info: dict) -> None:
    """Service trucks: replican el patron historico (Opening/Closing residuales)."""
    inflow = float(info.get("inflow", 0) or 0)
    outflow = float(info.get("outflow", 0) or 0)
    opening = outflow - inflow
    ws.cell(row=target_row, column=_COL_DATE,
            value=datetime.datetime.combine(month_first, datetime.time()))
    ws.cell(row=target_row, column=_COL_SITE, value=site_value)
    ws.cell(row=target_row, column=_COL_OPENING, value=opening)
    ws.cell(row=target_row, column=_COL_DELIVERIES, value=inflow)
    ws.cell(row=target_row, column=_COL_TRANSACTIONS, value=outflow)
    ws.cell(row=target_row, column=_COL_CALC_STOCK, value=opening)
    ws.cell(row=target_row, column=_COL_NET_CHANGE, value=0)
    ws.cell(row=target_row, column=_COL_CLOSING, value=opening)
    ws.cell(row=target_row, column=_COL_VARIANCE, value=0)
    ws.cell(row=target_row, column=_COL_PCT, value=0)
    ws.cell(row=target_row, column=_COL_TO_EQUIPMENT,
            value=info.get("to_equipment", outflow))
    ws.cell(row=target_row, column=_COL_OTHER,
            value=info.get("other_dispenses", 0))
    ws.cell(row=target_row, column=_COL_TRANSFERS,
            value=info.get("transfers_out", 0))


def _extend_tables_to_row(ws, target_row: int) -> None:
    try:
        tables = list(ws.tables.values())
    except AttributeError:
        return
    for table in tables:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        except Exception:
            continue
        if target_row > max_row:
            table.ref = "%s%d:%s%d" % (
                get_column_letter(min_col), min_row,
                get_column_letter(max_col), target_row)


def _to_excel_row_date(month_first: datetime.date) -> datetime.date:
    """Convencion Veridapt del Excel: la fila que contiene los datos del mes N
    se almacena con fecha = 1er dia del mes N+1.

    Asi, los datos de Abril 2026 (`month_first = 1/04/2026` desde el PDF) van
    a la fila con fecha `1/05/2026`. Al leer, history._previous_month_first
    revierte el desplazamiento y vuelve a mapear esa fila a Abril."""
    if month_first.month == 12:
        return month_first.replace(year=month_first.year + 1, month=1, day=1)
    return month_first.replace(month=month_first.month + 1, day=1)


def _write_one(ws, month_first: datetime.date, site_value: str,
               info: dict) -> tuple:
    # Aplicamos el +1 mes para respetar la convencion Veridapt del Excel.
    row_date = _to_excel_row_date(month_first)
    existing = _find_row_by_date(ws, row_date)
    if existing is not None:
        target_row = existing
        # Si la fila ya tiene datos y NO coinciden con `info`, marcamos como
        # "overwrite_diff" para que la UI lo destaque en el resumen.
        if _row_has_data(ws, target_row) and not _values_match(
                ws, target_row, info, site_value):
            action = "overwrite_diff"
        else:
            action = "updated"
    else:
        source_row = _last_data_row(ws)
        target_row = source_row + 1
        if source_row >= DATA_START:
            _copy_row_template(ws, source_row, target_row)
        _extend_tables_to_row(ws, target_row)
        action = "inserted"

    if info.get("is_truck") or str(site_value).startswith("TFL"):
        _write_one_truck(ws, target_row, row_date, site_value, info)
    else:
        _write_one_lfo(ws, target_row, row_date, site_value, info)
    return target_row, action


def _row_has_data(ws, row_idx: int) -> bool:
    """True si la fila tiene al menos Opening o Closing distintos de None."""
    op = ws.cell(row=row_idx, column=_COL_OPENING).value
    cl = ws.cell(row=row_idx, column=_COL_CLOSING).value
    return op is not None or cl is not None


def _values_match(ws, row_idx: int, info: dict, site_value: str,
                  tol: float = 1.0) -> bool:
    """True si la fila ya contiene los mismos numeros (Opening/Closing/Inflow)
    que `info`, dentro de la tolerancia. Usado para detectar la fila legacy
    que el bug grabo con datos de `info` en el slot equivocado."""
    def near(a, b):
        try:
            return abs(float(a or 0) - float(b or 0)) <= tol
        except (TypeError, ValueError):
            return False
    site_cell = ws.cell(row=row_idx, column=_COL_SITE).value
    if site_cell is not None and str(site_cell).strip().upper() != \
            str(site_value).strip().upper():
        return False
    return (near(ws.cell(row=row_idx, column=_COL_OPENING).value,
                 info.get("opening")) and
            near(ws.cell(row=row_idx, column=_COL_CLOSING).value,
                 info.get("closing")) and
            near(ws.cell(row=row_idx, column=_COL_DELIVERIES).value,
                 info.get("inflow")))


def detect_legacy_rows(excel_path: str, items: list) -> list:
    """Antes del fix de convencion, el cargador de PDF grababa los datos del
    mes N en la fila con fecha `1/N` (slot Veridapt del mes N-1). Esta funcion
    detecta esos casos: para cada (site, month_first, info) busca una fila en
    la fecha buggy (`month_first` sin +1 mes) cuyos valores coinciden con
    `info`. Devuelve una lista de dicts con la info necesaria para mostrar al
    usuario y, eventualmente, borrar la fila."""
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
    except Exception:
        return []
    out = []
    try:
        for site_key, month_first, info in items:
            target = SITE_TO_SHEET.get(site_key)
            if target is None:
                continue
            sheet_name, site_value = target
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            legacy = _find_row_by_date(ws, month_first)
            if legacy is None:
                continue
            if not _row_has_data(ws, legacy):
                continue
            if not _values_match(ws, legacy, info, site_value):
                continue
            out.append({
                "site_key": site_key,
                "sheet_name": sheet_name,
                "row": legacy,
                "legacy_date": month_first,
                "correct_date": _to_excel_row_date(month_first),
            })
    finally:
        wb.close()
    return out


def remove_rows(excel_path: str, items_to_remove: list) -> int:
    """Borra de Excel las filas indicadas. `items_to_remove` es una lista de
    dicts con al menos 'sheet_name' y 'row'. Devuelve cuantas filas se borraron.

    Borra de abajo hacia arriba para que los indices se mantengan validos."""
    if not items_to_remove:
        return 0
    _check_excel_not_locked(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    removed = 0
    try:
        by_sheet = {}
        for it in items_to_remove:
            by_sheet.setdefault(it["sheet_name"], []).append(int(it["row"]))
        for sheet_name, rows in by_sheet.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for r in sorted(set(rows), reverse=True):
                ws.delete_rows(r, 1)
                removed += 1
        wb.save(excel_path)
    finally:
        wb.close()
    return removed


def _check_excel_not_locked(excel_path: str) -> None:
    try:
        with open(excel_path, "a+b"):
            pass
    except PermissionError:
        raise PermissionError(
            "El archivo Excel parece estar abierto en Microsoft Excel:\n"
            "  %s\n\nCierre el archivo en Excel y vuelva a intentar."
            % excel_path)


def write_pdf_data(excel_path: str, site_key: str,
                   month_first: datetime.date, info: dict) -> dict:
    _check_excel_not_locked(excel_path)
    target = SITE_TO_SHEET.get(site_key)
    if target is None:
        return {"sheet": None, "row": None, "action": "unknown_site",
                "site": site_key}
    sheet_name, site_value = target

    wb = openpyxl.load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"sheet": None, "row": None, "action": "no_sheet",
                "site": site_key}
    ws = wb[sheet_name]
    row, action = _write_one(ws, month_first, site_value, info)
    wb.save(excel_path)
    wb.close()
    return {"sheet": sheet_name, "row": row, "action": action,
            "site": site_key}


def write_multiple(excel_path: str, items: list,
                   progress_cb=None) -> tuple:
    """items: lista de (site_key, month_first, info_dict).
    Devuelve (results, errors)."""
    _check_excel_not_locked(excel_path)
    results = []
    errors = []
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as exc:
        return [], [("(workbook)", str(exc))]

    total = len(items)
    for i, (site_key, month_first, info) in enumerate(items, start=1):
        if progress_cb:
            try:
                progress_cb(i, total, "Escribiendo %s..." % site_key)
            except Exception:
                pass
        try:
            target = SITE_TO_SHEET.get(site_key)
            if target is None:
                results.append((site_key, {"sheet": None, "row": None,
                                            "action": "unknown_site",
                                            "site": site_key}))
                continue
            sheet_name, site_value = target
            if sheet_name not in wb.sheetnames:
                results.append((site_key, {"sheet": None, "row": None,
                                            "action": "no_sheet",
                                            "site": site_key}))
                continue
            ws = wb[sheet_name]
            row, action = _write_one(ws, month_first, site_value, info)
            results.append((site_key, {"sheet": sheet_name, "row": row,
                                        "action": action, "site": site_key}))
        except Exception as exc:
            errors.append((site_key, str(exc)))

    try:
        wb.save(excel_path)
    except Exception as exc:
        errors.append(("(save)", str(exc)))
    wb.close()
    return results, errors
